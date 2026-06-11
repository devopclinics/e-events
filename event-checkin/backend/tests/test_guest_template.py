"""Guest-list template download + feature-aware import.

The template's columns are driven by the event's enabled add-ons, and the
importer must ingest exactly those columns — these tests cover the round trip.
"""
import io
import csv

import pytest
from sqlalchemy import select, func

from conftest import _Session
from app.models import Event, Guest, TicketType

BASE_COLS = ["first_name", "last_name", "email", "phone"]
SHIP_COLS = ["ship_address1", "ship_address2", "ship_city", "ship_state",
             "ship_postal", "ship_country"]


async def _enable(event_id, venue=False, logistics=False, ticket_names=()):
    async with _Session() as s:
        ev = await s.get(Event, event_id)
        ev.venue_access_enabled = venue
        ev.logistics_enabled = logistics
        for name in ticket_names:
            s.add(TicketType(event_id=event_id, name=name))
        await s.commit()


def _csv_header(resp):
    return next(csv.reader(io.StringIO(resp.text)))


async def _upload(ctx, event_id, text):
    return await ctx.client.post(
        f"/api/events/{event_id}/guests/upload",
        files={"file": ("guests.csv", text.encode(), "text/csv")},
    )


@pytest.mark.asyncio
async def test_template_base_columns(ctx):
    ctx.login(ctx.ids["user_a"])
    r = await ctx.client.get(f"/api/events/{ctx.ids['event_a']}/guests/template?fmt=csv")
    assert r.status_code == 200
    assert _csv_header(r) == BASE_COLS


@pytest.mark.asyncio
async def test_template_columns_follow_features(ctx):
    await _enable(ctx.ids["event_a"], venue=True, logistics=True, ticket_names=["VIP", "GA"])
    ctx.login(ctx.ids["user_a"])
    r = await ctx.client.get(f"/api/events/{ctx.ids['event_a']}/guests/template?fmt=csv")
    assert _csv_header(r) == BASE_COLS + ["ticket_type", "tags"] + SHIP_COLS
    # Sample row carries one of the event's real ticket types
    rows = list(csv.DictReader(io.StringIO(r.text)))
    assert rows[0]["ticket_type"] in ("VIP", "GA")


@pytest.mark.asyncio
async def test_template_xlsx(ctx):
    await _enable(ctx.ids["event_a"], venue=True, ticket_names=["VIP"])
    ctx.login(ctx.ids["user_a"])
    r = await ctx.client.get(f"/api/events/{ctx.ids['event_a']}/guests/template")
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert r.content[:4] == b"PK\x03\x04"


@pytest.mark.asyncio
async def test_template_cross_org_404(ctx):
    ctx.login(ctx.ids["user_b"])
    r = await ctx.client.get(f"/api/events/{ctx.ids['event_a']}/guests/template?fmt=csv")
    assert r.status_code == 404


@pytest.mark.asyncio
async def test_import_assigns_ticket_types_and_skips_sample(ctx):
    eid = ctx.ids["event_a"]
    await _enable(eid, venue=True, ticket_names=["VIP", "GA"])
    ctx.login(ctx.ids["user_a"])
    r = await _upload(ctx, eid, (
        "first_name,last_name,email,phone,ticket_type\n"
        "Jane,Doe,jane@example.com,+18325550100,VIP\n"   # template sample — must be skipped
        "Amy,Pond,amy@x.com,,vip\n"                       # case-insensitive match
        "Rory,Pond,rory@x.com,,Press\n"                   # unknown type — guest added, no ticket
    ))
    assert r.status_code == 200
    body = r.json()
    assert body["added"] == 2
    assert body["ticket_types_assigned"] == 1
    assert body["unknown_ticket_types"] == ["Press"]

    async with _Session() as s:
        amy = (await s.execute(select(Guest).where(Guest.email == "amy@x.com"))).scalar_one()
        rory = (await s.execute(select(Guest).where(Guest.email == "rory@x.com"))).scalar_one()
        vip = (await s.execute(select(TicketType).where(TicketType.name == "VIP"))).scalar_one()
        sample = (await s.execute(select(Guest).where(Guest.email == "jane@example.com"))).scalar_one_or_none()
    assert amy.ticket_type_id == vip.id
    assert rory.ticket_type_id is None
    assert sample is None


@pytest.mark.asyncio
async def test_reimport_backfills_ticket_and_address(ctx):
    eid = ctx.ids["event_a"]
    await _enable(eid, venue=True, logistics=True, ticket_names=["VIP"])
    ctx.login(ctx.ids["user_a"])
    # Seeded guest "G One <g@a.com>" exists with no ticket and no address.
    r = await _upload(ctx, eid, (
        "first_name,last_name,email,phone,ticket_type,ship_address1,ship_address2,"
        "ship_city,ship_state,ship_postal,ship_country\n"
        "G,One,g@a.com,,VIP,1 High St,,Geneva,GE,1201,Switzerland\n"
    ))
    body = r.json()
    assert body["added"] == 0 and body["skipped"] == 1
    assert body["ticket_types_assigned"] == 1
    assert body["addresses_added"] == 1

    async with _Session() as s:
        g = (await s.execute(select(Guest).where(Guest.email == "g@a.com"))).scalar_one()
    assert g.ticket_type_id is not None
    assert g.ship_city == "Geneva"


@pytest.mark.asyncio
async def test_headers_case_and_spacing_insensitive(ctx):
    eid = ctx.ids["event_a"]
    await _enable(eid, venue=True, ticket_names=["VIP"])
    ctx.login(ctx.ids["user_a"])
    r = await _upload(ctx, eid, (
        "First Name,LAST_NAME,Email,Phone,Ticket Type\n"
        "Clara,Oswald,clara@x.com,,VIP\n"
    ))
    body = r.json()
    assert body["added"] == 1
    assert body["ticket_types_assigned"] == 1

    async with _Session() as s:
        clara = (await s.execute(select(Guest).where(Guest.email == "clara@x.com"))).scalar_one()
    assert clara.first_name == "Clara" and clara.ticket_type_id is not None


@pytest.mark.asyncio
async def test_import_without_email_column(ctx):
    eid = ctx.ids["event_a"]
    ctx.login(ctx.ids["user_a"])
    csv_text = (
        "first_name,last_name,phone\n"
        "Donna,Noble,+18325550111\n"
        "Wilfred,Mott,\n"
    )
    r = await _upload(ctx, eid, csv_text)
    assert r.json()["added"] == 2
    # Re-import the same rows — must dedupe, not duplicate
    r = await _upload(ctx, eid, csv_text)
    body = r.json()
    assert body["added"] == 0 and body["skipped"] == 2

    async with _Session() as s:
        donna = (await s.execute(select(Guest).where(Guest.first_name == "Donna"))).scalar_one()
    assert donna.email is None and donna.phone == "+18325550111"


@pytest.mark.asyncio
async def test_import_respects_guest_cap(ctx):
    eid = ctx.ids["event_a"]  # free plan → FREE_GUEST_CAP=25; fixture seeds 1 guest
    ctx.login(ctx.ids["user_a"])
    rows = "\n".join(f"Guest,N{i},guest{i}@x.com," for i in range(30))
    r = await _upload(ctx, eid, "first_name,last_name,email,phone\n" + rows + "\n")
    body = r.json()
    assert body["added"] == 24
    assert body["over_cap"] == 6
    assert "Event Pass" in body["cap_note"]

    async with _Session() as s:
        n = await s.scalar(select(func.count(Guest.id)).where(Guest.event_id == eid))
    assert n == 25


@pytest.mark.asyncio
async def test_sample_rows_reported_separately(ctx):
    eid = ctx.ids["event_a"]
    ctx.login(ctx.ids["user_a"])
    r = await _upload(ctx, eid, (
        "first_name,last_name,email,phone\n"
        "Jane,Doe,jane@example.com,\n"
        "Amy,Pond,amy3@x.com,\n"
    ))
    body = r.json()
    assert body["added"] == 1
    assert body["sample_rows_skipped"] == 1


@pytest.mark.asyncio
async def test_xlsx_dropdown_uses_hidden_sheet(ctx):
    import openpyxl
    eid = ctx.ids["event_a"]
    await _enable(eid, venue=True, ticket_names=["VIP, Gold", "GA"])  # comma in name
    ctx.login(ctx.ids["user_a"])
    r = await ctx.client.get(f"/api/events/{eid}/guests/template")
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    assert "TicketTypes" in wb.sheetnames
    ref = wb["TicketTypes"]
    assert ref.sheet_state == "hidden"
    names = [row[0].value for row in ref.iter_rows()]
    assert "VIP, Gold" in names  # comma survives intact
    dvs = wb["Guests"].data_validations.dataValidation
    assert any("TicketTypes!" in (dv.formula1 or "") for dv in dvs)


@pytest.mark.asyncio
async def test_ticket_column_ignored_when_feature_off(ctx):
    eid = ctx.ids["event_a"]
    ctx.login(ctx.ids["user_a"])
    r = await _upload(ctx, eid, (
        "first_name,last_name,email,phone,ticket_type\n"
        "Amy,Pond,amy2@x.com,,VIP\n"
    ))
    body = r.json()
    assert body["added"] == 1
    assert "ticket_types_assigned" not in body
    assert "unknown_ticket_types" not in body
