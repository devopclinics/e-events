import csv
import html
import io
import re
import uuid
import httpx
from datetime import datetime
from urllib.parse import urlparse
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, BackgroundTasks, Response, Body
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import delete, select, func
from sqlalchemy.exc import IntegrityError
from ..database import get_db
from sqlalchemy import or_
from ..models import (
    ConsentSignature,
    EmailDeliveryEvent,
    Event,
    EventMessage,
    EventMessageDeliveryLog,
    EventMessageRead,
    EventMessageThread,
    ExperienceEvent,
    Guest,
    GuestExperienceProgress,
    GuestMenuChoice,
    GuestShipment,
    GuestTag,
    GuestTagLink,
    RSVPAnswer,
    RSVPQuestion,
    ScanEvent,
    SeatingTable,
    TableGroup,
    TicketType,
    User,
    EventUser,
    EventUserSection,
)
from ..schemas import GuestOut, GuestCreate, GuestUpdate, BulkAssignGroupRequest, ScanResult, WalkInRegister
from ..auth import require_event_admin, require_official
from ..entitlements import assert_within_guest_cap, guest_limit, can_use_paid_channels, last_credit_ledger_id, take_message_credit
from services.qr_service import generate_qr_bytes
from services.email_service import send_invite_email, send_manual_invite_email, send_simple_email
from ..template_resolve import load_overrides, channel_text as template_channel_text, email_override as template_email_override, channel_text_or_default as template_channel_or_default, email_or_default as template_email_or_default
from services.templates import TEMPLATE_DEFS, build_context as build_template_context
from .scanner import checkin_guard, perform_admission, queue_admission_email, queue_consent_copy_email
from services import messaging
from services.credit_ledger import send_with_credit_ledger
from ..services.experience import next_guest_steps, sync_guest_progress

router = APIRouter()

_E164_RE = re.compile(r'^\+[1-9]\d{6,14}$')
# Strip everything except digits and leading '+'.
_PHONE_STRIP = re.compile(r'[^\d+]')


def _normalize_phone(raw: str, default_country_code: str = "1") -> str | None:
    """Coerce a freeform phone string into E.164.

    Accepts: '+1 832-794-1707', '(832) 794-1707', '8327941707',
             '18327941707', '+18327941707'.
    Returns: '+18327941707' or None if the digits don't look like a valid number.

    `default_country_code` is used when the input has no '+' and looks like a
    local number. Defaults to US (1) — change if your event audience is elsewhere.
    """
    if not raw:
        return None
    cleaned = _PHONE_STRIP.sub('', raw.strip())
    if not cleaned:
        return None
    if cleaned.startswith('+'):
        candidate = cleaned
    elif cleaned.startswith('00'):
        # Some countries write international as 00<cc>... — treat as +<cc>...
        candidate = '+' + cleaned[2:]
    elif len(cleaned) == 10 and default_country_code == "1":
        # Bare US 10-digit number
        candidate = '+1' + cleaned
    elif len(cleaned) == 11 and cleaned.startswith('1') and default_country_code == "1":
        candidate = '+' + cleaned
    else:
        # Could be a longer non-US number missing its + — assume international and prepend.
        candidate = '+' + cleaned
    return candidate if _E164_RE.match(candidate) else None

# A real browser UA is required: SharePoint Online's abuse filter
# rejects requests with bot-looking User-Agents (returns 401/403).
_BROWSER_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
)


def _xlsx_to_csv_text(raw: bytes) -> str:
    """Convert xlsx binary to CSV-like text."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(400, "Excel file is empty")
    out = io.StringIO()
    writer = csv.writer(out)
    for row in rows:
        writer.writerow([str(c) if c is not None else "" for c in row])
    return out.getvalue()


def _decode_csv_bytes(raw: bytes, filename: str = "") -> str:
    name_lower = filename.lower()
    if name_lower.endswith(".xlsx") or raw[:4] == b"PK\x03\x04":
        return _xlsx_to_csv_text(raw)
    if name_lower.endswith(".xls") or raw[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        raise HTTPException(400, "Old .xls format is not supported — save as .xlsx or export as CSV")
    for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            return raw.decode(enc)
        except (UnicodeDecodeError, LookupError):
            continue
    raise HTTPException(400, "Could not decode file — save it as UTF-8 CSV and try again")


def _google_sheets_csv_url(url: str) -> str | None:
    """Convert a Google Sheets share URL to a CSV export URL."""
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", url)
    if not m:
        return None
    sheet_id = m.group(1)
    gid_m = re.search(r"[?&]gid=(\d+)", url)
    gid = gid_m.group(1) if gid_m else "0"
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"


def _google_sheets_xlsx_url(url: str) -> str | None:
    """Whole-workbook xlsx export URL for a Google Sheets share URL.

    Unlike format=csv, this also works for .xlsx files that were uploaded to
    Drive but never converted to a native Google Sheet (their share links carry
    rtpof=true). For those, format=csv returns HTTP 400, so this is the fallback.
    """
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", url)
    if not m:
        return None
    return f"https://docs.google.com/spreadsheets/d/{m.group(1)}/export?format=xlsx"


def _normalize_excel_url(url: str) -> str:
    """For OneDrive / SharePoint URLs, ensure ?download=1 is set so the
    redirect chain serves the raw file instead of the web viewer."""
    parsed = urlparse(url)
    host = parsed.netloc.lower()
    if not any(h in host for h in ("1drv.ms", "onedrive.live.com", "sharepoint.com")):
        return url
    if "download=1" in url:
        return url
    sep = "&" if parsed.query else "?"
    return f"{url}{sep}download=1"


async def _fetch_sheet_csv(url: str) -> tuple[bytes, str]:
    """Returns (raw_bytes, inferred_filename). Raises HTTPException(400) on failure."""
    google_url = _google_sheets_csv_url(url)
    fetch_url = google_url or _normalize_excel_url(url)

    try:
        async with httpx.AsyncClient(
            follow_redirects=True,
            timeout=30,
            headers={"User-Agent": _BROWSER_UA, "Accept": "*/*"},
        ) as client:
            resp = await client.get(fetch_url)
            # Uploaded-but-unconverted .xlsx files (share links carry rtpof=true)
            # reject format=csv with HTTP 400. Retry as whole-workbook xlsx, which
            # the _decode_csv_bytes pipeline handles transparently.
            if resp.status_code == 400 and google_url:
                xlsx_url = _google_sheets_xlsx_url(url)
                if xlsx_url:
                    resp = await client.get(xlsx_url)
    except Exception as e:
        raise HTTPException(400, f"Failed to reach spreadsheet URL: {e}")

    if resp.status_code >= 400:
        raise HTTPException(
            400,
            f"Could not fetch spreadsheet (HTTP {resp.status_code}). "
            "Make sure sharing is set to 'Anyone with the link can view' and "
            "paste the OneDrive/Excel Share → Copy link URL (not the browser address-bar URL).",
        )

    content_type = resp.headers.get("content-type", "").lower()
    body = resp.content
    if "text/html" in content_type or body.lstrip()[:200].lower().startswith(b"<!doctype html"):
        raise HTTPException(
            400,
            "The link returned an HTML page instead of a file — sharing is probably restricted. "
            "Open the file in OneDrive/Excel, click Share → 'Anyone with the link can view', "
            "copy that link, and paste it here.",
        )

    fname = "sheet.csv"
    if "spreadsheetml" in content_type or "xlsx" in content_type or body[:4] == b"PK\x03\x04":
        fname = "sheet.xlsx"
    return body, fname


# Guest shipping-address columns (logistics add-on) — template and import
# must agree on these names since they map 1:1 onto Guest attributes.
_SHIP_COLS = ("ship_address1", "ship_address2", "ship_city", "ship_state",
              "ship_postal", "ship_country")


def _template_columns(event: Event) -> list[str]:
    """Importable columns for this event, driven by its enabled add-ons.
    Must stay in lockstep with what _process_csv understands."""
    cols = ["first_name", "last_name", "email", "phone"]
    if event.venue_access_enabled:
        cols.append("ticket_type")
        cols.append("tags")
    if event.seating_enabled:
        cols.append("table_group")
    if event.logistics_enabled:
        cols.extend(_SHIP_COLS)
    return cols


# Sample row values. The example.com email doubles as the marker _process_csv
# uses to skip the row if the client forgets to delete it.
_TEMPLATE_SAMPLE = {
    "first_name": "Jane", "last_name": "Doe",
    "email": "jane@example.com", "phone": "+18325550100",
    "tags": "VIP; Press", "table_group": "VIP Tables",
    "ship_address1": "123 Main St", "ship_address2": "Apt 4B",
    "ship_city": "Houston", "ship_state": "TX",
    "ship_postal": "77002", "ship_country": "USA",
}


def _norm_header(h: str) -> str:
    """Normalize a CSV header: 'First Name', 'FIRST_NAME', 'first-name' →
    'first_name'. Lets clients use their own list without renaming columns."""
    return re.sub(r"[\s\-]+", "_", (h or "").strip().lower())


async def _process_csv(text: str, event_id: str, db: AsyncSession):
    text = text.replace('\r\n', '\n').replace('\r', '\n')
    reader = csv.DictReader(io.StringIO(text))
    fields = {_norm_header(f) for f in (reader.fieldnames or [])}
    missing = {'first_name', 'last_name'} - fields
    if missing:
        raise HTTPException(
            400,
            f"CSV is missing required columns: {', '.join(sorted(missing))}. "
            "Expected header: first_name,last_name,email,phone "
            "(case/spacing doesn't matter; email is optional)"
        )

    # Add-on columns (matching the downloadable template) are honored only when
    # the event has the feature enabled — extra columns are otherwise ignored.
    event = await db.get(Event, event_id)
    want_ticket = bool(event and event.venue_access_enabled) and "ticket_type" in fields
    want_shipping = bool(event and event.logistics_enabled) and "ship_address1" in fields
    want_tags = bool(event and event.venue_access_enabled) and "tags" in fields
    # Table-group assignment: any of these header aliases maps a guest to a group
    # (auto-creating the group if it doesn't exist yet).
    _GROUP_HEADERS = ("table_group", "table_tag", "assigned_table_tag")
    want_groups = bool(event and event.seating_enabled) and any(h in fields for h in _GROUP_HEADERS)
    group_id_by_key: dict[str, str] = {}        # tag/name (lower) → existing group id
    new_groups_by_key: dict[str, "TableGroup"] = {}  # created this run
    groups_assigned = groups_created = 0
    if want_groups:
        for grp in (await db.execute(select(TableGroup).where(TableGroup.event_id == event_id))).scalars():
            group_id_by_key[grp.tag.strip().lower()] = grp.id
            group_id_by_key.setdefault(grp.name.strip().lower(), grp.id)

    def _ensure_group_key(row: dict) -> str | None:
        """Return the lowercased group key for a row's table-group cell,
        auto-creating the group object on first sight. None when no cell set."""
        nonlocal groups_created
        cell = ""
        for h in _GROUP_HEADERS:
            cell = (row.get(h) or "").strip()
            if cell:
                break
        if not cell:
            return None
        key = cell.lower()
        if key not in group_id_by_key and key not in new_groups_by_key:
            grp = TableGroup(event_id=event_id, name=cell, tag=cell)
            db.add(grp)
            new_groups_by_key[key] = grp
            groups_created += 1
        return key

    pending_groups: list[tuple["Guest", str]] = []  # (guest, group key) resolved post-flush
    tt_by_name: dict[str, str] = {}
    if want_ticket:
        tts = (await db.execute(
            select(TicketType).where(TicketType.event_id == event_id)
        )).scalars().all()
        tt_by_name = {t.name.strip().lower(): t.id for t in tts}

    # Tag-based classification: resolve names to ids (auto-creating new tags),
    # and track existing links so re-imports don't duplicate.
    tag_id_by_name: dict[str, str] = {}            # existing tags → real id
    new_tag_by_name: dict[str, "GuestTag"] = {}    # created this run (id after flush)
    existing_links: set[tuple[str, str]] = set()
    pending_tags: list[tuple["Guest", list[str]]] = []   # (guest_obj, [name_key])
    tags_assigned = tags_created = 0
    if want_tags:
        for t in (await db.execute(select(GuestTag).where(GuestTag.event_id == event_id))).scalars():
            tag_id_by_name[t.name.strip().lower()] = t.id
        existing_links = set((await db.execute(
            select(GuestTagLink.guest_id, GuestTagLink.tag_id)
            .join(GuestTag, GuestTag.id == GuestTagLink.tag_id)
            .where(GuestTag.event_id == event_id))).all())

    def _resolve_tag_keys(cell: str) -> list[str]:
        """Return normalized tag name-keys for a row, auto-creating new tags
        (ids assigned at the later flush)."""
        nonlocal tags_created
        keys: list[str] = []
        for raw in re.split(r"[;,]", cell or ""):
            nm = raw.strip()
            if not nm:
                continue
            key = nm.lower()
            if key not in tag_id_by_name and key not in new_tag_by_name:
                tag = GuestTag(event_id=event_id, name=nm)
                db.add(tag)
                new_tag_by_name[key] = tag
                tags_created += 1
            if key not in keys:
                keys.append(key)
        return keys

    # Plan cap: bulk import must respect the same limit as single-add/RSVP.
    cap = guest_limit(event) if event else None
    current_count = 0
    if cap is not None:
        current_count = await db.scalar(
            select(func.count(Guest.id)).where(Guest.event_id == event_id)
        ) or 0

    added = skipped = invalid_phones = backfilled_phones = 0
    tickets_assigned = addresses_added = sample_rows = over_cap = 0
    unknown_tickets: set[str] = set()
    for raw_row in reader:
        row = {_norm_header(k): (v or "") for k, v in raw_row.items() if k is not None}
        email = row.get("email", "").strip().lower()
        first = row.get("first_name", "").strip()
        last = row.get("last_name", "").strip()
        if not first and not last:
            skipped += 1
            continue
        # The sample row shipped in the template uses a reserved example.com
        # address — skip it so an undeleted sample never becomes a guest.
        if email.endswith("@example.com"):
            sample_rows += 1
            continue

        phone_raw = row.get("phone", "").strip()
        normalized_phone: str | None = None
        if phone_raw:
            normalized_phone = _normalize_phone(phone_raw)
            if normalized_phone is None:
                invalid_phones += 1

        ticket_type_id: str | None = None
        if want_ticket:
            tt_name = (row.get("ticket_type") or "").strip()
            if tt_name:
                ticket_type_id = tt_by_name.get(tt_name.lower())
                if ticket_type_id is None:
                    unknown_tickets.add(tt_name)

        ship: dict[str, str | None] = {}
        if want_shipping and (row.get("ship_address1") or "").strip():
            ship = {c: (row.get(c) or "").strip() or None for c in _SHIP_COLS}

        row_tag_keys = _resolve_tag_keys(row.get("tags", "")) if want_tags else []
        row_group_key = _ensure_group_key(row) if want_groups else None

        # Deduplicate on first_name + last_name only (case-insensitive). Email is
        # NOT part of the key: spreadsheet re-syncs that change/add an email used
        # to create duplicate guests (count inflation). Email is treated as a
        # backfill field instead. (ported from prod)
        existing = (await db.execute(
            select(Guest).where(
                Guest.event_id == event_id,
                func.lower(Guest.first_name) == first.lower(),
                func.lower(Guest.last_name) == last.lower(),
            )
        )).scalars().first()
        if existing:
            # Backfill missing fields on re-import — useful when the source spreadsheet
            # now has an email/phone that was blank on the first import.
            if not existing.email and email:
                existing.email = email
            if not existing.phone and normalized_phone:
                existing.phone = normalized_phone
                backfilled_phones += 1
            if ticket_type_id and not existing.ticket_type_id:
                existing.ticket_type_id = ticket_type_id
                tickets_assigned += 1
            if ship and not existing.ship_address1:
                for c in _SHIP_COLS:
                    setattr(existing, c, ship[c])
                addresses_added += 1
            if row_tag_keys:
                pending_tags.append((existing, row_tag_keys))
            if row_group_key and not existing.assigned_table_group_id:
                pending_groups.append((existing, row_group_key))
            skipped += 1
            continue

        if cap is not None and current_count >= cap:
            over_cap += 1
            continue
        g = Guest(
            event_id=event_id,
            first_name=first,
            last_name=last,
            email=email or None,
            phone=normalized_phone,
            ticket_type_id=ticket_type_id,
            **ship,
        )
        db.add(g)
        added += 1
        current_count += 1
        if ticket_type_id:
            tickets_assigned += 1
        if ship:
            addresses_added += 1
        if row_tag_keys:
            pending_tags.append((g, row_tag_keys))
        if row_group_key:
            pending_groups.append((g, row_group_key))

    # Insert guests + any new tags first (assigns their ids), then the
    # guest↔tag links (FK-safe, deduped against pre-existing links).
    if want_tags:
        await db.flush()
        for guest, keys in pending_tags:
            for key in keys:
                tid = tag_id_by_name.get(key) or new_tag_by_name[key].id
                if (guest.id, tid) not in existing_links:
                    db.add(GuestTagLink(guest_id=guest.id, tag_id=tid))
                    existing_links.add((guest.id, tid))
                    tags_assigned += 1

    # Resolve table-group keys to ids (after flush so newly-created groups have
    # ids) and stamp each guest's single group FK.
    if want_groups:
        await db.flush()
        for guest, key in pending_groups:
            gid = group_id_by_key.get(key) or (new_groups_by_key[key].id if key in new_groups_by_key else None)
            if gid:
                guest.assigned_table_group_id = gid
                groups_assigned += 1

    await db.commit()
    result = {"added": added, "skipped": skipped}
    if backfilled_phones:
        result["backfilled_phones"] = backfilled_phones
    if invalid_phones:
        result["invalid_phones"] = invalid_phones
        result["phone_note"] = "Phones with invalid format were cleared (must be E.164 e.g. +447911123456)"
    if tickets_assigned:
        result["ticket_types_assigned"] = tickets_assigned
    if addresses_added:
        result["addresses_added"] = addresses_added
    if tags_assigned:
        result["tags_assigned"] = tags_assigned
    if tags_created:
        result["tags_created"] = tags_created
    if groups_assigned:
        result["table_groups_assigned"] = groups_assigned
    if groups_created:
        result["table_groups_created"] = groups_created
    if unknown_tickets:
        result["unknown_ticket_types"] = sorted(unknown_tickets)
        result["ticket_note"] = (
            "Unknown ticket types were ignored — create them in the Access tab "
            "(or download a fresh template), then re-import to assign."
        )
    if sample_rows:
        result["sample_rows_skipped"] = sample_rows
    if over_cap:
        result["over_cap"] = over_cap
        result["cap_note"] = (
            f"This event's plan allows up to {cap} guests — {over_cap} row(s) "
            "were not imported. Upgrade with an Event Pass to add more."
        )
    return result


def import_warning_summary(result: dict) -> str | None:
    """One-line human-readable warnings from a _process_csv result — shown in
    the sync-status UI so background sheet syncs don't swallow problems."""
    parts = []
    if result.get("over_cap"):
        parts.append(result["cap_note"])
    if result.get("unknown_ticket_types"):
        parts.append("Unknown ticket types ignored: " + ", ".join(result["unknown_ticket_types"]))
    if result.get("invalid_phones"):
        parts.append(f"{result['invalid_phones']} phone(s) had an invalid format and were left blank")
    return " · ".join(parts) or None


@router.get("/{event_id}/guests/template")
async def download_guest_template(event_id: str, fmt: str = "xlsx", db: AsyncSession = Depends(get_db), _: User = Depends(require_event_admin)):
    """Downloadable guest-list template. Columns reflect the event's enabled
    add-ons (ticket_type for venue access, ship_* for logistics), so what the
    client fills in is exactly what upload / URL import / live sync ingest."""
    event = await db.get(Event, event_id)
    if not event:
        raise HTTPException(404, "Event not found")
    if fmt not in ("csv", "xlsx"):
        raise HTTPException(400, "fmt must be csv or xlsx")

    cols = _template_columns(event)
    tt_names: list[str] = []
    if event.venue_access_enabled:
        tt_names = list((await db.execute(
            select(TicketType.name).where(TicketType.event_id == event_id)
            .order_by(TicketType.sort_order, TicketType.name)
        )).scalars())
    # Existing table-group names — offered as an Excel dropdown so clients pick a
    # real group instead of free-typing (new names still auto-create on import).
    tg_names: list[str] = []
    if event.seating_enabled:
        tg_names = list((await db.execute(
            select(TableGroup.name).where(TableGroup.event_id == event_id).order_by(TableGroup.name)
        )).scalars())
    sample = {**_TEMPLATE_SAMPLE,
              "ticket_type": tt_names[0] if tt_names else "",
              "table_group": tg_names[0] if tg_names else _TEMPLATE_SAMPLE.get("table_group", "")}
    sample_row = [sample.get(c, "") for c in cols]

    if fmt == "csv":
        out = io.StringIO()
        writer = csv.writer(out)
        writer.writerow(cols)
        writer.writerow(sample_row)
        return Response(
            out.getvalue(),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": 'attachment; filename="guest-template.csv"'},
        )

    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Guests"
    ws.append(cols)
    for i, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=i)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F766E")
        ws.column_dimensions[get_column_letter(i)].width = max(len(col) + 4, 16)
    ws.append(sample_row)
    for i in range(1, len(cols) + 1):
        ws.cell(row=2, column=i).font = Font(italic=True, color="94A3B8")
    ws.freeze_panes = "A2"

    if tt_names and "ticket_type" in cols:
        # Names live on a hidden sheet and the dropdown references the range —
        # unlike an inline list, this survives commas in names and any length.
        ref = wb.create_sheet("TicketTypes")
        for name in tt_names:
            ref.append([name])
        ref.sheet_state = "hidden"
        dv = DataValidation(
            type="list", formula1=f"=TicketTypes!$A$1:$A${len(tt_names)}", allow_blank=True,
        )
        dv.error = "Pick one of the event's ticket types"
        dv.prompt = "Ticket types: " + ", ".join(tt_names)
        ws.add_data_validation(dv)
        letter = get_column_letter(cols.index("ticket_type") + 1)
        dv.add(f"{letter}2:{letter}1001")

    if tg_names and "table_group" in cols:
        ref = wb.create_sheet("TableGroups")
        for name in tg_names:
            ref.append([name])
        ref.sheet_state = "hidden"
        dv = DataValidation(
            type="list", formula1=f"=TableGroups!$A$1:$A${len(tg_names)}", allow_blank=True,
        )
        # Not an error — clients may legitimately add a brand-new group, which
        # auto-creates on import. The dropdown is just a convenience.
        dv.showErrorMessage = False
        dv.prompt = "Existing groups: " + ", ".join(tg_names) + " (or type a new one)"
        ws.add_data_validation(dv)
        letter = get_column_letter(cols.index("table_group") + 1)
        dv.add(f"{letter}2:{letter}1001")

    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="guest-template.xlsx"'},
    )


@router.get("/{event_id}/guests/export")
async def export_guests(
    event_id: str,
    fmt: str = "csv",
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_event_admin),
):
    """Download the full guest list — including each guest's answers to the
    event's custom RSVP questions (one column per question). CSV or XLSX."""
    event = await db.get(Event, event_id)
    if not event:
        raise HTTPException(404, "Event not found")
    if fmt not in ("csv", "xlsx"):
        raise HTTPException(400, "fmt must be csv or xlsx")

    guests = (await db.execute(
        select(Guest).where(Guest.event_id == event_id).order_by(Guest.last_name, Guest.first_name)
    )).scalars().all()
    tnames = dict((await db.execute(
        select(SeatingTable.id, SeatingTable.name).where(SeatingTable.event_id == event_id))).all())
    gnames = dict((await db.execute(
        select(TableGroup.id, TableGroup.name).where(TableGroup.event_id == event_id))).all())
    ttnames = dict((await db.execute(
        select(TicketType.id, TicketType.name).where(TicketType.event_id == event_id))).all())
    questions = (await db.execute(
        select(RSVPQuestion).where(RSVPQuestion.event_id == event_id)
        .order_by(RSVPQuestion.sort_order, RSVPQuestion.question))).scalars().all()
    # One pass over all answers for this event → {(guest_id, question_id): answer}.
    answers: dict[tuple[str, str], str] = {}
    for gid, qid, ans in (await db.execute(
        select(RSVPAnswer.guest_id, RSVPAnswer.question_id, RSVPAnswer.answer)
        .join(Guest, Guest.id == RSVPAnswer.guest_id)
        .where(Guest.event_id == event_id)
    )).all():
        answers[(gid, qid)] = ans

    base_cols = ["First name", "Last name", "Email", "Phone", "RSVP status",
                 "Checked in", "Table", "Seat", "Group", "Ticket type", "VIP",
                 "Guest of", "Main guest ID", "Submitter email", "Submitter phone",
                 "Relationship", "Guest type", "RSVP notes"]
    cols = base_cols + [q.question for q in questions]

    def row_for(g: Guest) -> list[str]:
        row = [
            g.first_name or "", g.last_name or "", g.email or "", g.phone or "",
            g.rsvp_status or "", "Yes" if g.admitted else "No",
            tnames.get(g.table_id, "") if g.table_id else "",
            g.seat_number or "",
            gnames.get(g.assigned_table_group_id, "") if g.assigned_table_group_id else "",
            ttnames.get(g.ticket_type_id, "") if g.ticket_type_id else "",
            "Yes" if g.is_vip else "No",
            "Self / main invited guest" if g.rsvp_submitter_guest_id == g.id else (g.rsvp_submitter_name or ""),
            g.rsvp_submitter_guest_id or "",
            g.rsvp_submitter_email or "",
            g.rsvp_submitter_phone or "",
            g.rsvp_relationship or "",
            g.rsvp_guest_type or "",
            g.rsvp_notes or "",
        ]
        row += [answers.get((g.id, q.id), "") for q in questions]
        return row

    rows = [row_for(g) for g in guests]

    if fmt == "csv":
        out = io.StringIO()
        writer = csv.writer(out)
        writer.writerow(cols)
        writer.writerows(rows)
        return Response(
            out.getvalue(),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": 'attachment; filename="guest-list.csv"'},
        )

    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Guests"
    ws.append(cols)
    for i, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=i)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F766E")
        ws.column_dimensions[get_column_letter(i)].width = max(len(str(col)) + 2, 14)
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="guest-list.xlsx"'},
    )


@router.post("/{event_id}/guests/upload")
async def upload_guests(event_id: str, file: UploadFile = File(...), db: AsyncSession = Depends(get_db), _: User = Depends(require_event_admin)):
    event = await db.get(Event, event_id)
    if not event:
        raise HTTPException(404, "Event not found")

    raw = await file.read()
    text = _decode_csv_bytes(raw, file.filename or "")
    return await _process_csv(text, event_id, db)


def _dispatch_invite(background_tasks: BackgroundTasks, event: Event, guest: Guest,
                     overrides: dict | None = None,
                     rsvp_template_key: str = "rsvp_invitation") -> bool:
    """Fan out an invite across enabled channels for this event. Channel modules
    no-op silently when contact info is missing or creds aren't configured.

    In closed (invitation-only) mode we send the guest their personal RSVP link
    (/r/{invite_token}) rather than a ticket — the ticket QR is issued only after
    they confirm. In open mode we send the ticket directly, as before.

    `overrides` carries any customizable-template overrides for the event; when a
    channel has none, the default sender is used (unchanged behavior).

    Returns True if at least one channel was dispatched, False if the guest had no
    reachable channel (used to set invite_status sent/failed)."""
    if event.invite_mode == "closed":
        return _dispatch_rsvp_invite(background_tasks, event, guest, overrides, rsvp_template_key)

    ticket_url = f"{event.checkin_base_url.rstrip('/')}/scan/{guest.qr_token}"
    paid_channels = can_use_paid_channels(event)
    overrides = overrides or {}
    ctx = build_template_context(event, guest, extras={"ticket_link": ticket_url, "qr_code": ticket_url})
    dispatched = False

    if event.notify_email and guest.email:
        invite_key = "experience_invitation" if event.experience_enabled else "ticket_qr"
        ov = overrides.get(invite_key)
        spec = TEMPLATE_DEFS.get(invite_key, {})
        guest_data = {
            "guest_id": guest.id,
            "first_name": guest.first_name,
            "last_name":  guest.last_name,
            "email":      guest.email,
            "qr_token":   guest.qr_token,
            "event_id":   event.id,
            "message_kind": "experience_invitation" if event.experience_enabled else "invitation",
        }
        background_tasks.add_task(
            send_invite_email,
            guest_data, event.name, event.couples_name, event.checkin_base_url, event.event_date,
            event.seating_enabled, event.menu_enabled, event.partner_pairing_enabled,
            ov.subject if ov else spec.get("subject"), ov.email_body if ov else spec.get("email_body"),
            event.venue_name, event.venue_address, event.admission_note,
            event.invite_cover_image,
        )
        dispatched = True

    if paid_channels and event.notify_sms and guest.phone and guest.sms_consent and take_message_credit(event, "sms"):
        sms_text = (
            template_channel_or_default(overrides, "experience_invitation", "sms", ctx)
            if event.experience_enabled
            else template_channel_text(overrides, "sms_invitation", "sms", ctx)
        )
        if sms_text is not None:
            background_tasks.add_task(send_with_credit_ledger, last_credit_ledger_id(event), messaging.send_custom_sms, phone=guest.phone, body=sms_text)
        else:
            background_tasks.add_task(
                send_with_credit_ledger,
                last_credit_ledger_id(event),
                messaging.send_invite_sms,
                phone=guest.phone, first_name=guest.first_name,
                event_name=event.name, ticket_url=ticket_url, event_date=event.event_date,
            )
        dispatched = True

    if paid_channels and event.notify_whatsapp and guest.phone and guest.whatsapp_consent and take_message_credit(event, "whatsapp"):
        wa_text = (
            template_channel_or_default(overrides, "experience_invitation", "whatsapp", ctx)
            if event.experience_enabled
            else template_channel_text(overrides, "whatsapp_invitation", "whatsapp", ctx)
        )
        if wa_text is not None:
            background_tasks.add_task(send_with_credit_ledger, last_credit_ledger_id(event), messaging.send_custom_whatsapp, phone=guest.phone, body=wa_text)
        else:
            background_tasks.add_task(
                send_with_credit_ledger,
                last_credit_ledger_id(event),
                messaging.send_invite_whatsapp,
                phone=guest.phone, first_name=guest.first_name,
                event_name=event.name, ticket_url=ticket_url, event_date=event.event_date,
            )
        dispatched = True

    # MMS (ticket card) at invite time — super-admin-enabled per event.
    if (paid_channels and event.notify_mms and guest.phone and guest.sms_consent
            and messaging.mms_ready() and event.checkin_base_url and take_message_credit(event, "mms")):
        mms_key = "experience_invitation" if event.experience_enabled else "mms_invitation"
        mms_text = (template_channel_or_default(overrides, mms_key, "mms", ctx)
                    or f"Hi {guest.first_name}! You're invited to {event.name}.")
        card_url = f"{event.checkin_base_url.rstrip('/')}/api/scan/{guest.qr_token}/card.jpg"
        background_tasks.add_task(send_with_credit_ledger, last_credit_ledger_id(event), messaging.send_mms, phone=guest.phone, body=mms_text, media_url=card_url)
        dispatched = True

    return dispatched


def _dispatch_rsvp_invite(background_tasks: BackgroundTasks, event: Event, guest: Guest,
                          overrides: dict | None = None,
                          template_key: str = "rsvp_invitation") -> bool:
    """Closed-mode invite: send the guest their personal RSVP link. Generates a
    one-per-guest invite_token on first send (mutation persisted by the caller's
    commit). No ticket/QR yet — that's issued when they confirm. Returns True if
    at least one channel fired."""
    if not guest.invite_token:
        guest.invite_token = str(uuid.uuid4())
    invite_url = f"{event.checkin_base_url.rstrip('/')}/r/{guest.invite_token}"
    name = f"{guest.first_name} {guest.last_name}".strip() or "Guest"
    paid_channels = can_use_paid_channels(event)
    overrides = overrides or {}
    ctx = build_template_context(event, guest, extras={"rsvp_link": invite_url})
    dispatched = False

    if event.notify_email and guest.email:
        if template_key == "rsvp_invitation":
            subj, body = template_email_override(overrides, template_key, ctx)
        else:
            subj, body = template_email_or_default(overrides, template_key, ctx)
        if body is not None:
            background_tasks.add_task(
                send_simple_email,
                guest.email,
                subj or f"You're invited — {event.name}",
                body,
                event.id,
                None,
                guest.id,
                template_key,
            )
        else:
            background_tasks.add_task(
                send_manual_invite_email,
                name=name, email=guest.email, invite_url=invite_url,
                event_name=event.name, event_date=event.event_date,
                invite_message=event.invite_message,
                event_id=event.id,
                guest_id=guest.id,
            )
        dispatched = True

    if paid_channels and event.notify_sms and guest.phone and guest.sms_consent and take_message_credit(event, "sms"):
        if template_key == "rsvp_invitation":
            sms_text = template_channel_text(overrides, template_key, "sms", ctx)
        else:
            sms_text = template_channel_or_default(overrides, template_key, "sms", ctx)
        if sms_text is not None:
            background_tasks.add_task(send_with_credit_ledger, last_credit_ledger_id(event), messaging.send_custom_sms, phone=guest.phone, body=sms_text)
        else:
            background_tasks.add_task(
                send_with_credit_ledger,
                last_credit_ledger_id(event),
                messaging.send_manual_invite_sms,
                phone=guest.phone, name=name,
                event_name=event.name, invite_url=invite_url,
            )
        dispatched = True

    if paid_channels and event.notify_whatsapp and guest.phone and guest.whatsapp_consent and take_message_credit(event, "whatsapp"):
        wa_text = template_channel_text(overrides, template_key, "whatsapp", ctx)
        if wa_text is not None:
            background_tasks.add_task(send_with_credit_ledger, last_credit_ledger_id(event), messaging.send_custom_whatsapp, phone=guest.phone, body=wa_text)
        elif template_key == "rsvp_reminder":
            background_tasks.add_task(
                send_with_credit_ledger,
                last_credit_ledger_id(event),
                messaging.send_rsvp_reminder_whatsapp,
                phone=guest.phone, first_name=guest.first_name,
                event_name=event.name, invite_url=invite_url,
            )
        else:
            background_tasks.add_task(
                send_with_credit_ledger,
                last_credit_ledger_id(event),
                messaging.send_manual_invite_whatsapp,
                phone=guest.phone, name=name,
                event_name=event.name, invite_url=invite_url,
            )
        dispatched = True

    return dispatched


def dispatch_approval_accepted(background_tasks: BackgroundTasks, event: Event, guest: Guest,
                               overrides: dict | None = None) -> bool:
    """Send approval-accepted notices with the guest's issued ticket link."""
    ticket_url = f"{event.checkin_base_url.rstrip('/')}/scan/{guest.qr_token}"
    overrides = overrides or {}
    ctx = build_template_context(event, guest, extras={"ticket_link": ticket_url, "qr_code": ticket_url})
    sent = False

    if event.notify_email and guest.email:
        ov = overrides.get("approval_accepted")
        spec = TEMPLATE_DEFS.get("approval_accepted", {})
        subject_tmpl = (ov.subject if ov and ov.subject else None) or spec.get("subject")
        body_tmpl = (ov.email_body if ov and ov.email_body else None) or spec.get("email_body")
        if body_tmpl:
            if "{{qr_code" not in body_tmpl:
                body_tmpl = f"{body_tmpl}<p>{{{{qr_code}}}}</p>"
            if "{{ticket_link" not in body_tmpl:
                body_tmpl = f'{body_tmpl}<p><a href="{{{{ticket_link}}}}">View your pass</a></p>'
            background_tasks.add_task(
                send_invite_email,
                {
                    "guest_id": guest.id,
                    "first_name": guest.first_name,
                    "last_name": guest.last_name,
                    "email": guest.email,
                    "qr_token": guest.qr_token,
                    "event_id": event.id,
                    "message_kind": "approval_accepted",
                },
                event.name,
                event.couples_name,
                event.checkin_base_url,
                event.event_date,
                event.seating_enabled,
                event.menu_enabled,
                event.partner_pairing_enabled,
                subject_tmpl,
                body_tmpl,
                event.venue_name,
                event.venue_address,
                event.admission_note,
                event.invite_cover_image,
            )
            sent = True

    if (can_use_paid_channels(event) and event.notify_sms and guest.phone
            and guest.sms_consent and take_message_credit(event, "sms")):
        sms = template_channel_or_default(overrides, "approval_accepted", "sms", ctx)
        if sms:
            background_tasks.add_task(send_with_credit_ledger, last_credit_ledger_id(event), messaging.send_custom_sms, phone=guest.phone, body=sms)
            sent = True

    if (can_use_paid_channels(event) and event.notify_whatsapp and guest.phone
            and guest.whatsapp_consent and take_message_credit(event, "whatsapp")):
        wa_text = template_channel_text(overrides, "approval_accepted", "whatsapp", ctx)
        if wa_text is not None:
            background_tasks.add_task(send_with_credit_ledger, last_credit_ledger_id(event), messaging.send_custom_whatsapp, phone=guest.phone, body=wa_text)
        else:
            background_tasks.add_task(
                send_with_credit_ledger,
                last_credit_ledger_id(event),
                messaging.send_approval_accepted_whatsapp,
                phone=guest.phone, first_name=guest.first_name,
                event_name=event.name, ticket_url=ticket_url,
            )
        sent = True

    return sent


def dispatch_simple_notice(background_tasks: BackgroundTasks, event: Event, guest: Guest,
                           key: str, overrides: dict, extras: dict | None = None) -> bool:
    """Send a simple notification template to a guest — used for
    decline/rejected confirmations. Gated by the event's channel flags; uses the
    event override or the registry default. Returns True if anything was scheduled."""
    ctx = build_template_context(event, guest, extras=extras or {})
    sent = False
    if event.notify_email and guest.email:
        subj, body = template_email_or_default(overrides, key, ctx)
        if body:
            background_tasks.add_task(send_simple_email, guest.email, subj or event.name, body, event.id, None, guest.id, key)
            sent = True
    if (can_use_paid_channels(event) and event.notify_sms and guest.phone
            and guest.sms_consent and take_message_credit(event, "sms")):
        sms = template_channel_or_default(overrides, key, "sms", ctx)
        if sms:
            background_tasks.add_task(send_with_credit_ledger, last_credit_ledger_id(event), messaging.send_custom_sms, phone=guest.phone, body=sms)
            sent = True
    wa = template_channel_text(overrides, key, "whatsapp", ctx)
    if (wa is not None or key in {"rsvp_decline", "approval_pending", "rsvp_confirmation", "approval_rejected"}) and (
            can_use_paid_channels(event) and event.notify_whatsapp and guest.phone
            and guest.whatsapp_consent and take_message_credit(event, "whatsapp")):
        if wa is not None:
            background_tasks.add_task(send_with_credit_ledger, last_credit_ledger_id(event), messaging.send_custom_whatsapp, phone=guest.phone, body=wa)
            sent = True
        elif key == "rsvp_decline":
            background_tasks.add_task(
                send_with_credit_ledger,
                last_credit_ledger_id(event),
                messaging.send_rsvp_decline_whatsapp,
                phone=guest.phone, first_name=guest.first_name,
                event_name=event.name,
            )
            sent = True
        elif key == "approval_pending":
            background_tasks.add_task(
                send_with_credit_ledger,
                last_credit_ledger_id(event),
                messaging.send_approval_pending_whatsapp,
                phone=guest.phone, first_name=guest.first_name,
                event_name=event.name,
            )
            sent = True
        elif key == "rsvp_confirmation":
            background_tasks.add_task(
                send_with_credit_ledger,
                last_credit_ledger_id(event),
                messaging.send_rsvp_confirmation_whatsapp,
                phone=guest.phone, first_name=guest.first_name,
                event_name=event.name, event_date=event.event_date,
            )
            sent = True
        elif key == "approval_rejected":
            background_tasks.add_task(
                messaging.send_approval_rejected_whatsapp,
                phone=guest.phone, first_name=guest.first_name,
                event_name=event.name,
            )
            sent = True
    return sent


async def _dispatch_experience_next_steps(
    background_tasks: BackgroundTasks,
    event: Event,
    guest: Guest,
    db: AsyncSession,
    overrides: dict | None = None,
) -> bool:
    """Send a guest their currently actionable Experience steps."""
    if not event.notify_email or not guest.email:
        return False

    def session_text(session: dict | None) -> str:
        if not isinstance(session, dict):
            return ""
        parts = [str(session[key]) for key in ("topic", "date") if session.get(key)]
        times = " - ".join(str(session.get(key)) for key in ("start_time", "end_time") if session.get(key))
        if times:
            parts.append(times)
        if session.get("room"):
            parts.append(str(session["room"]))
        if session.get("speaker"):
            parts.append(f"Speaker: {session['speaker']}")
        return " · ".join(parts)

    rows = await next_guest_steps(event.id, guest.id, db)
    step_items = []
    ticket_url = f"{event.checkin_base_url.rstrip('/')}/scan/{guest.qr_token}" if event.checkin_base_url else ""
    for step, _progress in rows:
        config = step.config or {}
        messages = config.get("messages") if isinstance(config.get("messages"), dict) else {}
        description = (messages.get("guest") or config.get("guest_message") or step.description or "").strip()
        session = session_text(config.get("session"))
        if session:
            description = f"{description}\n{session}" if description else session
        action = (
            f'<br><a href="{html.escape(ticket_url + "#consent", quote=True)}">Open consent form</a>'
            if step.type == "consent" and ticket_url else ""
        )
        step_items.append({
            "title": step.title,
            "description": description,
            "required": bool(step.required),
            "action": action,
        })
    if step_items:
        list_items = "".join(
            "<li><strong>{title}</strong>{required}{description}{action}</li>".format(
                title=html.escape(item["title"]),
                required=" <span>(required)</span>" if item["required"] else "",
                description=f"<br>{html.escape(item['description'])}" if item["description"] else "",
                action=item.get("action") or "",
            )
            for item in step_items
        )
        steps_html = f"<ol>{list_items}</ol>"
        steps_text = "; ".join(
            f"{item['title']}{' (required)' if item['required'] else ''}" for item in step_items
        )
    else:
        steps_html = "<p>You have no pending Experience steps right now.</p>"
        steps_text = "No pending steps right now."

    ctx = build_template_context(
        event,
        guest,
        extras={
            "ticket_link": ticket_url,
            "qr_code": ticket_url,
            "experience_steps": steps_html,
            "experience_steps_text": steps_text,
        },
    )
    subj, body = template_email_or_default(overrides or {}, "experience_next_steps", ctx)
    if not body:
        return False
    background_tasks.add_task(send_simple_email, guest.email, subj or f"Your next steps — {event.name}", body, event.id, None, guest.id, "experience_next_steps")
    return True


async def import_from_source_url(url: str, event_id: str, db: AsyncSession) -> dict:
    """Fetch a Google Sheets / OneDrive / Excel Online URL and merge rows into the event.
    Shared by the manual import endpoint and the background sync poller."""
    raw, fname = await _fetch_sheet_csv(url)
    text = _decode_csv_bytes(raw, fname)
    return await _process_csv(text, event_id, db)


@router.post("/{event_id}/guests/import-url")
async def import_guests_from_url(
    event_id: str,
    body: dict = Body(...),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_event_admin),
):
    event = await db.get(Event, event_id)
    if not event:
        raise HTTPException(404, "Event not found")

    url = (body.get("url") or "").strip()
    if not url:
        raise HTTPException(400, "url is required")

    return await import_from_source_url(url, event_id, db)


@router.post("/{event_id}/guests", response_model=GuestOut, status_code=201)
async def add_guest(event_id: str, data: GuestCreate, db: AsyncSession = Depends(get_db), _: User = Depends(require_event_admin)):
    """Manually add a single guest — used for VVIP walk-ins not on the original list.
    Email is optional; phone validated if provided. No invite is auto-sent."""
    event = await db.get(Event, event_id)
    if not event:
        raise HTTPException(404, "Event not found")

    first = data.first_name.strip()
    last = data.last_name.strip()
    if not first or not last:
        raise HTTPException(400, "first_name and last_name are required")
    email = (data.email or "").strip().lower()
    phone_raw = (data.phone or "").strip()
    phone = _normalize_phone(phone_raw) if phone_raw else None
    if phone_raw and phone is None:
        raise HTTPException(400, "Phone format not recognised. Use E.164 (e.g. +18327941707) or US 10-digit.")

    count = await db.scalar(select(func.count()).where(Guest.event_id == event_id)) or 0
    assert_within_guest_cap(event, count)

    group_id = None
    if data.assigned_table_group_id:
        grp = await db.get(TableGroup, data.assigned_table_group_id)
        if not grp or grp.event_id != event_id:
            raise HTTPException(404, "Table group not found for this event")
        group_id = grp.id

    guest = Guest(event_id=event_id, first_name=first, last_name=last, email=email,
                  phone=phone, is_vip=bool(data.is_vip), assigned_table_group_id=group_id,
                  is_walk_in=bool(data.is_walk_in))
    db.add(guest)
    await db.commit()
    await db.refresh(guest)
    return guest


@router.get("/{event_id}/guests", response_model=list[GuestOut])
async def list_guests(event_id: str, db: AsyncSession = Depends(get_db), _: User = Depends(require_event_admin)):
    result = await db.execute(
        select(Guest).where(Guest.event_id == event_id).order_by(Guest.last_name, Guest.first_name)
    )
    guests = result.scalars().all()
    # Decorate with the (non-mapped) table-group name so the guest list / export
    # can show it without an extra round-trip.
    names = dict((await db.execute(
        select(TableGroup.id, TableGroup.name).where(TableGroup.event_id == event_id)
    )).all())
    for g in guests:
        g.table_group_name = names.get(g.assigned_table_group_id)
    guest_ids = [g.id for g in guests]
    if guest_ids:
        rows = (await db.execute(
            select(EmailDeliveryEvent)
            .where(EmailDeliveryEvent.event_id == event_id, EmailDeliveryEvent.guest_id.in_(guest_ids))
            .order_by(EmailDeliveryEvent.occurred_at.desc(), EmailDeliveryEvent.created_at.desc())
        )).scalars().all()
        latest_by_guest = {}
        for row in rows:
            latest_by_guest.setdefault(row.guest_id, row)
        for g in guests:
            row = latest_by_guest.get(g.id)
            if row:
                g.email_delivery_status = row.status
                g.email_delivery_event_type = row.event_type
                g.email_delivery_kind = row.message_kind
                g.email_delivery_at = row.occurred_at
    return guests


@router.get("/{event_id}/guests/{guest_id}/rsvp-answers")
async def guest_rsvp_answers(
    event_id: str,
    guest_id: str,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_event_admin),
):
    """The guest's answers to the event's custom RSVP questions, ordered as the
    questions appear on the invite page. Read-only — for the organizer to review
    in the guest detail panel."""
    guest = await db.get(Guest, guest_id)
    if not guest or guest.event_id != event_id:
        raise HTTPException(404, "Guest not found")
    rows = (await db.execute(
        select(RSVPQuestion.question, RSVPQuestion.question_type, RSVPAnswer.answer)
        .join(RSVPAnswer, RSVPAnswer.question_id == RSVPQuestion.id)
        .where(RSVPAnswer.guest_id == guest_id, RSVPQuestion.event_id == event_id)
        .order_by(RSVPQuestion.sort_order, RSVPQuestion.question)
    )).all()
    return [{"question": q, "question_type": qt, "answer": a} for q, qt, a in rows]


@router.post("/{event_id}/guests/bulk-assign-group")
async def bulk_assign_table_group(
    event_id: str,
    body: BulkAssignGroupRequest,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_event_admin),
):
    """Assign (or clear, when table_group_id is null) a table group for one or
    many guests. Used by the guest profile and the Guests-tab bulk action."""
    if not await db.get(Event, event_id):
        raise HTTPException(404, "Event not found")
    if body.table_group_id is not None:
        grp = await db.get(TableGroup, body.table_group_id)
        if not grp or grp.event_id != event_id:
            raise HTTPException(404, "Table group not found for this event")
    updated = 0
    for gid in body.guest_ids:
        guest = await db.get(Guest, gid)
        if not guest or guest.event_id != event_id:
            continue
        guest.assigned_table_group_id = body.table_group_id
        updated += 1
    await db.commit()
    return {"ok": True, "updated": updated, "table_group_id": body.table_group_id}


@router.patch("/{event_id}/guests/{guest_id}", response_model=GuestOut)
async def update_guest(
    event_id: str,
    guest_id: str,
    data: GuestUpdate,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_event_admin),
):
    """Edit a guest's core fields from the admin guest-edit modal (ported from prod)."""
    guest = await db.get(Guest, guest_id)
    if not guest or guest.event_id != event_id:
        raise HTTPException(404, "Guest not found")
    if data.first_name is not None:
        guest.first_name = data.first_name.strip()
    if data.last_name is not None:
        guest.last_name = data.last_name.strip()
    if data.email is not None:
        guest.email = data.email.strip() or None
    if data.phone is not None:
        phone = _normalize_phone(data.phone.strip()) if data.phone.strip() else None
        if data.phone.strip() and phone is None:
            raise HTTPException(400, "Invalid phone format — use E.164 e.g. +447911123456")
        guest.phone = phone
    if data.is_vip is not None:
        guest.is_vip = data.is_vip
    if data.sms_consent is not None:
        guest.sms_consent = data.sms_consent
    if data.whatsapp_consent is not None:
        guest.whatsapp_consent = data.whatsapp_consent

    # Manual table/seat assignment. "" clears; a value validates the table and
    # guards against double-booking the same seat (409).
    if data.table_id is not None:
        new_table_id = data.table_id.strip() or None
        if new_table_id:
            table = await db.get(SeatingTable, new_table_id)
            if not table or table.event_id != event_id:
                raise HTTPException(404, "Table not found for this event")
            # Capacity guard: don't over-fill a table. Only enforced when moving
            # the guest onto a different table; counts other guests already there
            # (incl. table-only assignments the seat-clash check below can't see).
            if guest.table_id != new_table_id:
                others = await db.scalar(
                    select(func.count(Guest.id)).where(
                        Guest.table_id == new_table_id, Guest.id != guest.id
                    )
                ) or 0
                if others >= table.capacity:
                    raise HTTPException(409, f"{table.name} is full (capacity {table.capacity}).")
    # Resolve the intended final (table, seat) so the clash check runs on the
    # would-be values BEFORE we mutate `guest` — otherwise the query autoflushes
    # our pending change into the unique index and raises there.
    final_table = new_table_id if data.table_id is not None else guest.table_id
    if data.table_id is not None and new_table_id is None:
        final_seat = None  # no table → no seat
    elif data.seat_number is not None:
        final_seat = data.seat_number.strip() or None
    else:
        final_seat = guest.seat_number

    # Seat must be a real seat on the table (1..capacity). Bounding to capacity
    # is what makes concurrent manual assignment overflow-proof: only `capacity`
    # distinct seats can exist, and the unique index rejects any duplicate — so
    # two admins can't both land a guest on the same full table via different
    # out-of-range seat numbers.
    if final_table and final_seat:
        seat_table = await db.get(SeatingTable, final_table)
        if seat_table and (not str(final_seat).isdigit() or not (1 <= int(final_seat) <= seat_table.capacity)):
            raise HTTPException(400, f"Seat number must be between 1 and {seat_table.capacity} for {seat_table.name}.")

    # Reject a seat already held by another guest on the same table.
    if final_table and final_seat:
        clash = await db.scalar(
            select(Guest).where(
                Guest.event_id == event_id,
                Guest.table_id == final_table,
                Guest.seat_number == final_seat,
                Guest.id != guest.id,
            )
        )
        if clash:
            raise HTTPException(409, f"Seat {final_seat} on that table is already taken by {clash.first_name} {clash.last_name}.")

    if data.table_id is not None:
        guest.table_id = new_table_id
        if new_table_id is None:
            guest.seat_number = None  # no table → no seat
    if data.seat_number is not None:
        guest.seat_number = data.seat_number.strip() or None

    # Unique index backstop for a concurrent seat grab between check and commit.
    try:
        await db.commit()
    except IntegrityError:
        await db.rollback()
        raise HTTPException(409, f"Seat {guest.seat_number} on that table was just taken — refresh and pick another seat.")
    await sync_guest_progress(event_id, guest.id, db, source="admin")
    await db.commit()
    await db.refresh(guest)
    return guest


# ── Manual check-in (no QR) ─────────────────────────────────────────────────────

def _mask_phone(phone: str | None) -> str | None:
    """Show only the last 4 digits, mask the rest (privacy on the search list)."""
    if not phone:
        return None
    digits = "".join(ch for ch in phone if ch.isdigit())
    if len(digits) <= 4:
        return phone
    return "•" * (len(digits) - 4) + digits[-4:]


@router.get("/{event_id}/guests/search")
async def search_guests(
    event_id: str,
    q: str = "",
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_official),
):
    """Manual check-in search: partial, case-insensitive match across first name,
    last name (incl. full name) and phone, all at once. Gated to events with
    manual check-in on and to staff assigned to the event."""
    event = await db.get(Event, event_id)
    if not event:
        raise HTTPException(404, "Event not found")
    if not event.manual_checkin_enabled:
        raise HTTPException(403, "Manual check-in is not enabled for this event")
    blocked = await checkin_guard(event, current_user, db)
    if blocked:
        raise HTTPException(403, blocked.message)

    term = (q or "").strip().lower()
    if len(term) < 2:
        return []
    p = f"%{term}%"
    rows = (await db.execute(
        select(Guest).where(
            Guest.event_id == event_id,
            or_(
                func.lower(Guest.first_name).like(p),
                func.lower(Guest.last_name).like(p),
                func.lower(Guest.first_name + " " + Guest.last_name).like(p),
                func.lower(func.coalesce(Guest.phone, "")).like(p),
            ),
        ).order_by(Guest.admitted, Guest.last_name, Guest.first_name).limit(25)
    )).scalars().all()

    table_ids = {g.table_id for g in rows if g.table_id}
    names: dict[str, str] = {}
    if table_ids:
        names = dict((await db.execute(
            select(SeatingTable.id, SeatingTable.name).where(SeatingTable.id.in_(table_ids))
        )).all())

    return [{
        "id": g.id,
        "first_name": g.first_name,
        "last_name": g.last_name,
        "full_name": f"{g.first_name} {g.last_name}".strip(),
        "phone_masked": _mask_phone(g.phone),
        "table_name": names.get(g.table_id),
        "seat_number": g.seat_number,
        "is_vip": g.is_vip,
        "admitted": g.admitted,
        "admitted_at": g.admitted_at.isoformat() if g.admitted_at else None,
        "rsvp_status": g.rsvp_status,
    } for g in rows]


async def _allowed_section_ids(event_id, user, db) -> set[str] | None:
    """The table-group ids this user may route door check-ins into, or None when
    unrestricted (all sections). Unrestricted = the member has no per-section rows,
    which also covers admins/superadmins who have no EventUser row for the event."""
    eu = await db.scalar(select(EventUser).where(EventUser.event_id == event_id, EventUser.user_id == user.id))
    if not eu:
        return None
    ids = set((await db.execute(
        select(EventUserSection.table_group_id).where(EventUserSection.event_user_id == eu.id)
    )).scalars())
    return ids or None


async def _member_section(event, user, section_id, db) -> str | None:
    """Resolve which section (table group) a door check-in routes to under section
    mode, honoring the staffer's assignment:
      - a chosen section is validated and must be one they're allowed;
      - with no choice, auto-route only when they're restricted to exactly one.
    """
    allowed = await _allowed_section_ids(event.id, user, db)
    if section_id:
        if allowed is not None and section_id not in allowed:
            raise HTTPException(403, "You are not assigned to this section")
        grp = await db.get(TableGroup, section_id)
        if not grp or grp.event_id != event.id:
            raise HTTPException(404, "Section not found for this event")
        return section_id
    if allowed is not None and len(allowed) == 1:
        return next(iter(allowed))
    return None


async def _resolve_section_group(event, user, section_id, db) -> str | None:
    """Walk-in routing. Section mode ON → the staffer's assigned/active section;
    OFF → the event's single walk_in_table_group_id (unchanged legacy behavior)."""
    if event.section_mode_enabled:
        return await _member_section(event, user, section_id, db)
    return event.walk_in_table_group_id or None


@router.post("/{event_id}/guests/walk-in", response_model=ScanResult)
async def register_walk_in(
    event_id: str,
    body: WalkInRegister,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_official),
):
    """Register a walk-in guest at the door: create the guest, auto-assign them to
    a table group, and admit — all in one step (no QR). When section mode is on the
    guest is routed to the scanner's active section (body.table_group_id); otherwise
    to the event's single walk_in_table_group_id."""
    event = await db.get(Event, event_id)
    if not event:
        raise HTTPException(404, "Event not found")
    if not (event.walk_in_enabled or event.manual_checkin_enabled):
        raise HTTPException(403, "Walk-in registration is not enabled for this event")
    blocked = await checkin_guard(event, current_user, db)
    if blocked:
        return blocked
    first = (body.first_name or "").strip()
    if not first:
        raise HTTPException(400, "Guest name is required")
    group_id = await _resolve_section_group(event, current_user, body.table_group_id, db)
    phone = _normalize_phone(body.phone.strip()) if (body.phone or "").strip() else None
    guest = Guest(
        event_id=event_id, first_name=first, last_name=(body.last_name or "").strip(),
        phone=phone, qr_generated_at=datetime.utcnow(),
        assigned_table_group_id=group_id,
        is_walk_in=True,
    )
    db.add(guest)
    await db.flush()
    return await perform_admission(guest, event, background_tasks, db)


@router.post("/{event_id}/guests/{guest_id}/checkin", response_model=ScanResult)
async def manual_checkin(
    event_id: str,
    guest_id: str,
    background_tasks: BackgroundTasks,
    table_group_id: str | None = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_official),
):
    """Admit a guest by id (manual check-in) — runs the exact same admission flow
    as a QR scan (seat assignment, notifications, SSE broadcast).

    Section mode: if the device passes its active section (table_group_id) and the
    guest has NO table group yet, route them into that section. A guest who already
    has a group keeps it — the section never overrides an existing assignment."""
    event = await db.get(Event, event_id)
    if not event:
        raise HTTPException(404, "Event not found")
    if not event.manual_checkin_enabled:
        raise HTTPException(403, "Manual check-in is not enabled for this event")
    guest = await db.get(Guest, guest_id)
    if not guest or guest.event_id != event_id:
        return ScanResult(status="invalid", message="Guest not found for this event.")
    blocked = await checkin_guard(event, current_user, db)
    if blocked:
        return blocked
    if event.section_mode_enabled and not guest.assigned_table_group_id:
        sec = await _member_section(event, current_user, table_group_id, db)
        if sec:
            guest.assigned_table_group_id = sec
    return await perform_admission(guest, event, background_tasks, db)


@router.get("/{event_id}/my-sections")
async def my_sections(
    event_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_official),
):
    """Sections (table groups) the signed-in staffer may check guests into on the
    scanner. Restricted members get only their assigned sections; everyone else
    (unassigned/admin) gets all of the event's groups. One result → auto-route on
    the device; two or more → the device shows a picker limited to these."""
    event = await db.get(Event, event_id)
    if not event:
        raise HTTPException(404, "Event not found")
    groups = (await db.execute(
        select(TableGroup).where(TableGroup.event_id == event_id)
        .order_by(TableGroup.sort_order, TableGroup.name)
    )).scalars().all()
    allowed = await _allowed_section_ids(event_id, current_user, db)
    if allowed is not None:
        groups = [g for g in groups if g.id in allowed]
    return {
        "section_mode_enabled": bool(event.section_mode_enabled),
        "sections": [{"id": g.id, "name": g.name} for g in groups],
    }


@router.delete("/{event_id}/guests/{guest_id}", status_code=204)
async def delete_guest(event_id: str, guest_id: str, db: AsyncSession = Depends(get_db), _: User = Depends(require_event_admin)):
    guest = await db.get(Guest, guest_id)
    if not guest or guest.event_id != event_id:
        raise HTTPException(404, "Guest not found")
    await db.execute(
        Guest.__table__.update()
        .where(Guest.event_id == event_id, Guest.partner_guest_id == guest_id)
        .values(partner_guest_id=None)
    )
    await db.execute(delete(GuestExperienceProgress).where(GuestExperienceProgress.guest_id == guest_id))
    await db.execute(delete(ExperienceEvent).where(ExperienceEvent.guest_id == guest_id))
    await db.execute(delete(ConsentSignature).where(ConsentSignature.guest_id == guest_id))
    await db.execute(delete(ScanEvent).where(ScanEvent.guest_id == guest_id))
    await db.execute(delete(GuestTagLink).where(GuestTagLink.guest_id == guest_id))
    await db.execute(delete(GuestShipment).where(GuestShipment.guest_id == guest_id))
    await db.execute(delete(GuestMenuChoice).where(GuestMenuChoice.guest_id == guest_id))
    await db.execute(delete(RSVPAnswer).where(RSVPAnswer.guest_id == guest_id))
    await db.execute(delete(EventMessageRead).where(EventMessageRead.guest_id == guest_id))
    await db.execute(delete(EventMessageDeliveryLog).where(EventMessageDeliveryLog.guest_id == guest_id))
    await db.execute(delete(EventMessage).where(EventMessage.guest_id == guest_id))
    await db.execute(delete(EventMessageThread).where(EventMessageThread.guest_id == guest_id))
    await db.delete(guest)
    await db.commit()


@router.post("/{event_id}/guests/generate-qr")
async def generate_qr_codes(event_id: str, db: AsyncSession = Depends(get_db), _: User = Depends(require_event_admin)):
    event = await db.get(Event, event_id)
    if not event:
        raise HTTPException(404, "Event not found")

    result = await db.execute(
        select(Guest).where(Guest.event_id == event_id, Guest.qr_generated_at == None)
    )
    guests = result.scalars().all()

    for guest in guests:
        guest.qr_generated_at = datetime.utcnow()

    await db.commit()
    return {"generated": len(guests)}


@router.post("/{event_id}/guests/send-invites")
async def send_invites(event_id: str, background_tasks: BackgroundTasks, db: AsyncSession = Depends(get_db), _: User = Depends(require_event_admin)):
    event = await db.get(Event, event_id)
    if not event:
        raise HTTPException(404, "Event not found")

    result = await db.execute(
        select(Guest).where(
            Guest.event_id == event_id,
            Guest.qr_generated_at != None,
            Guest.invite_sent_at == None,
        )
    )
    guests = result.scalars().all()

    overrides = await load_overrides(event_id, db)
    for guest in guests:
        ok = _dispatch_invite(background_tasks, event, guest, overrides)
        guest.invite_sent_at = datetime.utcnow()
        guest.invite_status = "sent" if ok else "failed"

    await db.commit()
    return {"queued": len(guests)}


@router.post("/{event_id}/guests/send-batch")
async def send_invites_batch(
    event_id: str,
    body: dict = Body(default={}),
    background_tasks: BackgroundTasks = None,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_event_admin),
):
    """Send (or resend) invites in bulk.

    Body:
      guest_ids: list[str] | None  — if omitted, applies to all guests in the event
      force: bool                  — if true, re-send even to already-invited guests
                                     (otherwise only unsent guests get the email)
    """
    event = await db.get(Event, event_id)
    if not event:
        raise HTTPException(404, "Event not found")

    guest_ids = body.get("guest_ids") or []
    force = bool(body.get("force", False))

    q = select(Guest).where(Guest.event_id == event_id)
    if guest_ids:
        q = q.where(Guest.id.in_(guest_ids))
    if not force:
        q = q.where(Guest.invite_sent_at == None)  # noqa: E711

    guests = (await db.execute(q)).scalars().all()

    queued = 0
    now = datetime.utcnow()
    overrides = await load_overrides(event_id, db)
    for guest in guests:
        # Auto-generate QR timestamp on first send so it can also be a no-op for never-touched guests.
        if not guest.qr_generated_at:
            guest.qr_generated_at = now
        rsvp_template_key = (
            "rsvp_reminder"
            if event.invite_mode == "closed" and force and guest.invite_sent_at and guest.rsvp_status == "invited"
            else "rsvp_invitation"
        )
        ok = _dispatch_invite(background_tasks, event, guest, overrides, rsvp_template_key)
        guest.invite_sent_at = now
        guest.invite_status = "sent" if ok else "failed"
        queued += 1

    await db.commit()
    return {"queued": queued, "force": force, "scope": "selected" if guest_ids else "all"}


@router.post("/{event_id}/guests/{guest_id}/resend-invite")
async def resend_invite(event_id: str, guest_id: str, background_tasks: BackgroundTasks, db: AsyncSession = Depends(get_db), _: User = Depends(require_event_admin)):
    event = await db.get(Event, event_id)
    if not event:
        raise HTTPException(404, "Event not found")
    guest = await db.get(Guest, guest_id)
    if not guest or guest.event_id != event_id:
        raise HTTPException(404, "Guest not found")
    # Open mode emails a ticket (needs a QR); closed mode emails an RSVP link (no QR yet).
    if event.invite_mode != "closed" and not guest.qr_generated_at:
        raise HTTPException(400, "Generate QR codes first before sending invites")

    ok = _dispatch_invite(background_tasks, event, guest, await load_overrides(event_id, db))
    guest.invite_sent_at = datetime.utcnow()
    guest.invite_status = "sent" if ok else "failed"
    await db.commit()
    return {"ok": True}


@router.post("/{event_id}/guests/{guest_id}/resend-email")
async def resend_guest_email(
    event_id: str,
    guest_id: str,
    background_tasks: BackgroundTasks,
    body: dict = Body(...),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_event_admin),
):
    """Resend a specific guest-facing email from the portal."""
    kind = (body.get("kind") or "").strip()
    allowed = {"invitation", "admission", "experience_next_steps", "consent_copy"}
    if kind not in allowed:
        raise HTTPException(400, f"kind must be one of: {', '.join(sorted(allowed))}")

    event = await db.get(Event, event_id)
    if not event:
        raise HTTPException(404, "Event not found")
    guest = await db.get(Guest, guest_id)
    if not guest or guest.event_id != event_id:
        raise HTTPException(404, "Guest not found")
    if not guest.email:
        raise HTTPException(400, "Guest does not have an email address")

    overrides = await load_overrides(event_id, db)
    if kind == "invitation":
        if event.invite_mode != "closed" and not guest.qr_generated_at:
            raise HTTPException(400, "Generate QR codes first before sending invites")
        ok = _dispatch_invite(background_tasks, event, guest, overrides)
        guest.invite_sent_at = datetime.utcnow()
        guest.invite_status = "sent" if ok else "failed"
        await db.commit()
        return {"ok": ok, "kind": kind}

    if kind == "admission":
        if not guest.admitted:
            raise HTTPException(400, "Guest has not been admitted yet")
        ok = await queue_admission_email(background_tasks, event, guest, db)
        return {"ok": ok, "kind": kind}

    if kind == "consent_copy":
        ok = await queue_consent_copy_email(background_tasks, event, guest, db)
        return {"ok": ok, "kind": kind}

    ok = await _dispatch_experience_next_steps(background_tasks, event, guest, db, overrides)
    return {"ok": ok, "kind": kind}


@router.post("/{event_id}/guests/{guest_id}/approve")
async def approve_rsvp(event_id: str, guest_id: str, background_tasks: BackgroundTasks, db: AsyncSession = Depends(get_db), _: User = Depends(require_event_admin)):
    """Approve a pending self-registered RSVP: confirm the guest, issue their QR,
    and send the ticket. No-op-safe if the guest is already confirmed."""
    event = await db.get(Event, event_id)
    if not event:
        raise HTTPException(404, "Event not found")
    guest = await db.get(Guest, guest_id)
    if not guest or guest.event_id != event_id:
        raise HTTPException(404, "Guest not found")

    now = datetime.utcnow()
    guest.rsvp_status = "confirmed"
    if not guest.qr_generated_at:
        guest.qr_generated_at = now
    guest.invite_sent_at = now
    overrides = await load_overrides(event_id, db)
    ok = dispatch_approval_accepted(background_tasks, event, guest, overrides)
    if not ok:
        ok = _dispatch_invite(background_tasks, event, guest, overrides)
    guest.invite_status = "sent" if ok else "failed"
    await db.commit()
    return {"ok": True, "rsvp_status": "confirmed"}


@router.post("/{event_id}/guests/{guest_id}/reject")
async def reject_rsvp(event_id: str, guest_id: str, background_tasks: BackgroundTasks, db: AsyncSession = Depends(get_db), _: User = Depends(require_event_admin)):
    """Reject a pending RSVP — marks the guest declined (no ticket) and notifies
    them via the 'approval_rejected' template. Keeps the record for history."""
    event = await db.get(Event, event_id)
    guest = await db.get(Guest, guest_id)
    if not guest or guest.event_id != event_id:
        raise HTTPException(404, "Guest not found")
    guest.rsvp_status = "declined"
    guest.rsvp_responded_at = datetime.utcnow()
    if event and event.notify_rsvp_responses:
        dispatch_simple_notice(background_tasks, event, guest, "approval_rejected",
                               await load_overrides(event_id, db))
    await db.commit()
    return {"ok": True, "rsvp_status": "declined"}


@router.post("/{event_id}/guests/{guest_id}/invite-token")
async def ensure_invite_token(event_id: str, guest_id: str, db: AsyncSession = Depends(get_db), _: User = Depends(require_event_admin)):
    """Mint (or return) a guest's personal RSVP-link token so the planner can
    copy /r/{token} from the dashboard without having to send the invite first."""
    guest = await db.get(Guest, guest_id)
    if not guest or guest.event_id != event_id:
        raise HTTPException(404, "Guest not found")
    if not guest.invite_token:
        guest.invite_token = str(uuid.uuid4())
        await db.commit()
        await db.refresh(guest)
    event = await db.get(Event, event_id)
    return {
        "invite_token": guest.invite_token,
        "invite_url": f"{event.checkin_base_url.rstrip('/')}/r/{guest.invite_token}",
    }


@router.get("/{event_id}/guests/{guest_id}/qr.png")
async def get_guest_qr(event_id: str, guest_id: str, db: AsyncSession = Depends(get_db)):
    guest = await db.get(Guest, guest_id)
    if not guest or guest.event_id != event_id:
        raise HTTPException(404, "Guest not found")
    event = await db.get(Event, event_id)
    qr_bytes = generate_qr_bytes(guest.qr_token, event.checkin_base_url)
    return Response(content=qr_bytes, media_type="image/png")
